import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import os
import sys

def extract_text_from_pdf(file_path):
    """
    Mengekstrak teks dari file PDF dengan pdfplumber untuk hasil yang lebih akurat.
    """
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            extracted_text = page.extract_text()
            if extracted_text:
                text += extracted_text + "\n"
    return text

def process_data(text):
    """
    Memproses teks dan mengekstrak informasi yang diperlukan.
    """
    lines = text.split("\n")
    
    slot_pattern = r"Default Slot (\d+)"
    variant_pattern = r"Variant: ([^\n]+)"
    sku_pattern = r"SKU: ([^\n]+)"
    qty_pattern = r"Qty: (\d+)"
    
    ignore_keywords = ["Jumlah produk", "Buyer Notes", "Palembang", "TANJUNG PURA JAKARTA BARAT", "Tanggal Cetak", "Dicetak Oleh", "Jumlah Pesanan", "Picking List", "PICK", "Bogor Loji", "Halaman"]
    
    data = []
    current_product = {}
    nama_produk_buffer = []
    
    for line in lines:
        if any(keyword in line for keyword in ignore_keywords):
            continue
        
        if "Default Slot" in line:
            if current_product:
                if nama_produk_buffer:
                    current_product["Nama Produk"] = " ".join(nama_produk_buffer).strip()
                    nama_produk_buffer = []
                data.append(current_product)
            current_product = {}
            
            slot_match = re.search(slot_pattern, line)
            if slot_match:
                current_product["Slot"] = slot_match.group(1)
        
        variant_match = re.search(variant_pattern, line)
        if variant_match:
            current_product["Variant"] = variant_match.group(1)
        
        sku_match = re.search(sku_pattern, line)
        if sku_match:
            current_product["SKU"] = sku_match.group(1)
        
        qty_match = re.search(qty_pattern, line)
        if qty_match:
            current_product["Qty"] = qty_match.group(1)
        
        if not any(keyword in line for keyword in ["Default Slot", "Variant", "SKU", "Qty"]):
            nama_produk_buffer.append(line.strip())
    
    if current_product:
        if nama_produk_buffer:
            current_product["Nama Produk"] = " ".join(nama_produk_buffer).strip()
        data.append(current_product)
    
    return data

def clean_data(data):
    """
    Membersihkan data dari spasi berlebihan atau karakter yang tidak diperlukan.
    """
    for item in data:
        for key, value in item.items():
            if isinstance(value, str):
                item[key] = value.strip()
    return data

def save_to_excel(data, output_file):
    """
    Menyimpan data ke dalam file Excel dengan wrap text, border, header bold, dan middle align.
    """
    wb = Workbook()
    ws = wb.active
    
    headers = ["Nama Produk", "Variant", "SKU", "Slot"]
    ws.append(headers)
    
    # Membuat header bold dan middle align
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for item in data:
        row = [
            str(item.get("Nama Produk", "")),
            str(item.get("Variant", "")),
            str(item.get("SKU", "")),
            str(item.get("Slot", "")),
        ]
        ws.append(row)
    
    # Mengatur lebar kolom agar lebih nyaman dibaca
    ws.column_dimensions['A'].width = 50  # Nama Produk lebih panjang
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 4
    
    #Membuat cell di A3
    max_row = ws.max_row  # Ambil baris terakhir yang ada di worksheet
    ws.move_range(f"A3:A{max_row}", rows=1, cols=0)  # Geser hanya sampai baris terakhir
    ws["A3"] = " "  # Tambahkan isi baru ke A3
    
    # Menerapkan Wrap Text dan Middle Align untuk seluruh sel kecuali header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
    
    # Loop dari baris terakhir ke atas agar penghapusan tidak mengubah indeks sebelumnya
    for row in range(ws.max_row, 1, -1):  
        non_empty_cells = [cell for cell in ws[row] if cell.value is not None and str(cell.value).strip() != ""]

        if len(non_empty_cells) <= 1:  # Jika hanya 0 atau 1 sel yang terisi
            ws.delete_rows(row)  # Hapus baris tersebut
    
    # Menambahkan Border ke seluruh sel
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    wb.save(output_file)

def main(file_path):
    """
    Fungsi utama untuk memproses file PDF.
    """
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    output_file = os.path.join(os.path.dirname(file_path), f"{file_name}.xlsx")
    
    text = extract_text_from_pdf(file_path)
    data = process_data(text)
    cleaned_data = clean_data(data)
    save_to_excel(cleaned_data, output_file)
    
    print(f"Data telah disimpan ke {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Penggunaan: python rapihkan_data_pdf_to_excel.py <file_pdf>")
        sys.exit(1)
    
    pdf_file_path = sys.argv[1]
    main(pdf_file_path)
